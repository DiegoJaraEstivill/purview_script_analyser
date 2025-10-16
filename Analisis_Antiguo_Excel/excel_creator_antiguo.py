from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def crear_nombre_archivo():
    """Crea el nombre del archivo Excel con formato PurviewInf_DDMMAAAA_HHMM"""
    ahora = datetime.now()
    fecha = ahora.strftime("%d%m%Y")  # día-mes-año
    hora = ahora.strftime("%H%M")     # hora-minuto
    return f"PurviewInf_{fecha}_{hora}.xlsx"

def crear_excel_purview(datos_registros):
    """
    Crea un archivo Excel con los datos de los registros de Purview
    
    Args:
        datos_registros (list): Lista de diccionarios con los datos de cada registro
        
    Returns:
        str: Nombre del archivo creado
    """
    # Crear workbook y worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos Purview"
    
    # Definir headers base (5 originales)
    headers_base = ['RecordID', 'Creation Date', 'Record Type', 'Operation', 'User ID']
    
    # Definir headers de auditoría (39 campos)
    headers_audit = [
        'Creation Time', 'Audit ID', 'Audit Operation', 'Organization ID', 'Audit Record Type',
        'User Key', 'User Type', 'Version', 'Workload', 'Client IP', 'Audit User ID',
        'Authentication Type', 'Browser Name', 'Browser Version', 'Correlation ID',
        'Event Source', 'Geo Location', 'Is Managed Device', 'Item Type', 'List ID',
        'List Item Unique ID', 'Platform', 'Site', 'User Agent', 'Web ID',
        'Device Display Name', 'Event Signature', 'Machine ID', 'File Sync Bytes Committed',
        'High Priority Media Processing', 'Implicit Share', 'List Base Type', 'List Server Template',
        'Source Relative URL', 'Source File Name', 'Source File Extension',
        'Application Display Name', 'Site URL', 'Object ID'
    ]
    
    # Combinar todos los headers (5 + 39 = 44 columnas)
    headers = headers_base + headers_audit
    
    # Agregar headers con formato
    for col, header in enumerate(headers, 1):
        celda = ws.cell(row=1, column=col, value=header)
        celda.font = Font(bold=True, color='FFFFFF')
        celda.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    # Agregar datos fila por fila
    for fila_num, registro in enumerate(datos_registros, 2):  # Empezar en fila 2 (después del header)
        # Agregar datos base (columnas 1-5)
        ws.cell(row=fila_num, column=1, value=registro.get('record_id', 'N/A'))
        ws.cell(row=fila_num, column=2, value=registro.get('creation_date', 'N/A'))
        ws.cell(row=fila_num, column=3, value=registro.get('record_type', 'N/A'))
        ws.cell(row=fila_num, column=4, value=registro.get('operation', 'N/A'))
        ws.cell(row=fila_num, column=5, value=registro.get('user_id', 'N/A'))
        
        # Agregar datos de auditoría (columnas 6-44) de forma dinámica
        audit_keys = [
            'creation_time', 'audit_id', 'audit_operation', 'organization_id', 'audit_record_type',
            'user_key', 'user_type', 'version', 'workload', 'client_ip', 'audit_user_id',
            'authentication_type', 'browser_name', 'browser_version', 'correlation_id',
            'event_source', 'geo_location', 'is_managed_device', 'item_type', 'list_id',
            'list_item_unique_id', 'platform', 'site', 'user_agent', 'web_id',
            'device_display_name', 'event_signature', 'machine_id', 'file_sync_bytes_committed',
            'high_priority_media_processing', 'implicit_share', 'list_base_type', 'list_server_template',
            'source_relative_url', 'source_file_name', 'source_file_extension',
            'application_display_name', 'site_url', 'object_id'
        ]
        
        for i, key in enumerate(audit_keys, 6):  # Empezar en columna 6
            ws.cell(row=fila_num, column=i, value=registro.get(key, 'N/A'))
    
    # Ajustar ancho de columnas automáticamente
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        # Establecer un ancho mínimo de 12 y máximo de 50
        adjusted_width = min(max(max_length + 2, 12), 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar archivo con nombre generado automáticamente
    nombre_archivo = crear_nombre_archivo()
    wb.save(nombre_archivo)
    
    print(f"✅ Archivo Excel creado exitosamente: {nombre_archivo}")
    print(f"📊 Total de registros procesados: {len(datos_registros)}")
    
    return nombre_archivo