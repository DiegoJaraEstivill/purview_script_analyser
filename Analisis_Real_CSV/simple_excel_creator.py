from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def crear_nombre_archivo():
    """Crea el nombre del archivo Excel con formato PurviewInf_DDMMAAAA_HHMM"""
    ahora = datetime.now()
    fecha = ahora.strftime("%d%m%Y")  # día-mes-año
    hora = ahora.strftime("%H%M")     # hora-minuto
    return f"PurviewInf_{fecha}_{hora}.xlsx"

def crear_excel_simple(datos_registros):
    """
    Crea un archivo Excel con solo los primeros 5 campos
    
    Args:
        datos_registros (list): Lista de diccionarios con los datos de cada registro
        
    Returns:
        str: Nombre del archivo creado
    """
    print("\n📊 Creando archivo Excel...")
    print("=" * 60)
    
    # Crear workbook y worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos Purview"
    
    # Definir headers (solo 5 columnas)
    headers = ['RecordId', 'CreationDate', 'RecordType', 'Operation', 'UserId']
    
    # Agregar headers con formato
    for col, header in enumerate(headers, 1):
        celda = ws.cell(row=1, column=col, value=header)
        celda.font = Font(bold=True, color='FFFFFF', size=12)
        celda.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        celda.alignment = Alignment(horizontal='center', vertical='center')
    
    # Agregar datos fila por fila
    for fila_num, registro in enumerate(datos_registros, 2):  # Empezar en fila 2
        ws.cell(row=fila_num, column=1, value=registro.get('RecordId', 'N/A'))
        ws.cell(row=fila_num, column=2, value=registro.get('CreationDate', 'N/A'))
        ws.cell(row=fila_num, column=3, value=registro.get('RecordType', 'N/A'))
        ws.cell(row=fila_num, column=4, value=registro.get('Operation', 'N/A'))
        ws.cell(row=fila_num, column=5, value=registro.get('UserId', 'N/A'))
    
    # Ajustar ancho de columnas automáticamente
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        # Establecer un ancho mínimo de 15 y máximo de 50
        adjusted_width = min(max(max_length + 2, 15), 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar archivo con nombre generado automáticamente
    nombre_archivo = crear_nombre_archivo()
    wb.save(nombre_archivo)
    
    print(f"✅ Archivo Excel creado exitosamente: {nombre_archivo}")
    print(f"📊 Total de registros procesados: {len(datos_registros)}")
    print(f"📋 Columnas creadas: {len(headers)}")
    print("=" * 60)
    
    return nombre_archivo

