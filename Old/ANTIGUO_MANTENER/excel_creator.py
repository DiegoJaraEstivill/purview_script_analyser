from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def crear_nombre_archivo():
    """Crea el nombre del archivo Excel con formato PurviewInf_DDMMAAAA_HHMM"""
    ahora = datetime.now()
    fecha = ahora.strftime("%d%m%Y")  # día-mes-año
    hora = ahora.strftime("%H%M")     # hora-minuto
    return f"PurviewInf_{fecha}_{hora}.xlsx"

def crear_excel_purview_completo(datos_registros):
    """
    FUNCIÓN ANTIGUA - DEPRECADA - 44 columnas con JSON
    Esta función está archivada. Usar crear_excel_purview() en su lugar
    """
    raise DeprecationWarning("Esta función está deprecada. Usar crear_excel_purview() en su lugar")

def crear_excel_purview(datos_registros):
    """
    Crea un archivo Excel con solo los primeros 5 campos básicos
    
    Args:
        datos_registros (list): Lista de diccionarios con los datos de cada registro
        
    Returns:
        str: Nombre del archivo creado
    """
    # Crear workbook y worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos Purview"
    
    # Definir headers (solo 5 columnas básicas)
    headers = ['RecordId', 'CreationDate', 'RecordType', 'Operation', 'UserId']
    
    # Agregar headers con formato
    for col, header in enumerate(headers, 1):
        celda = ws.cell(row=1, column=col, value=header)
        celda.font = Font(bold=True, color='FFFFFF')
        celda.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    # Agregar datos fila por fila - SOLO 5 columnas
    for fila_num, registro in enumerate(datos_registros, 2):  # Empezar en fila 2 (después del header)
        ws.cell(row=fila_num, column=1, value=registro.get('RecordId', 'N/A'))
        ws.cell(row=fila_num, column=2, value=registro.get('CreationDate', 'N/A'))
        ws.cell(row=fila_num, column=3, value=registro.get('RecordType', 'N/A'))
        ws.cell(row=fila_num, column=4, value=registro.get('Operation', 'N/A'))
        ws.cell(row=fila_num, column=5, value=registro.get('UserId', 'N/A'))
    
    # Ajustar ancho de columnas automáticamente
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        # Establecer un ancho mínimo de 12 y máximo de 50
        adjusted_width = min(max(max_length + 2, 12), 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar archivo con nombre generado automáticamente
    nombre_archivo = crear_nombre_archivo()
    wb.save(nombre_archivo)
    
    print(f"✅ Archivo Excel creado exitosamente: {nombre_archivo}")
    print(f"📊 Total de registros procesados: {len(datos_registros)}")
    
    return nombre_archivo