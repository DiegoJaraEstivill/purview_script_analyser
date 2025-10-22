from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

def crear_nombre_archivo():
    """Crea el nombre del archivo Excel con formato PurviewInf_DDMMAAAA_HHMM"""
    ahora = datetime.now()
    fecha = ahora.strftime("%d%m%Y")  # día-mes-año
    hora = ahora.strftime("%H%M")     # hora-minuto
    return f"PurviewInf_Completo_{fecha}_{hora}.xlsx"

def crear_excel_dos_pestanas(datos_usuarios_normales, campos_usuarios_normales, 
                            datos_usuarios_sharepoint, campos_usuarios_sharepoint):
    """
    Crea un archivo Excel con DOS PESTANAS separadas por tipo de usuario
    
    Args:
        datos_usuarios_normales (list): Datos de usuarios normales
        campos_usuarios_normales (list): Campos de usuarios normales
        datos_usuarios_sharepoint (list): Datos de usuarios SharePoint system
        campos_usuarios_sharepoint (list): Campos de usuarios SharePoint system
        
    Returns:
        str: Nombre del archivo creado
    """
    print("\nCREANDO ARCHIVO EXCEL CON DOS PESTANAS...")
    print("=" * 80)
    
    # Crear workbook
    wb = Workbook()
    
    # Eliminar la hoja por defecto
    wb.remove(wb.active)
    
    # Crear pestaña para usuarios normales
    ws_normales = wb.create_sheet("Usuarios Normales")
    crear_pestana_usuarios_normales(ws_normales, datos_usuarios_normales, campos_usuarios_normales)
    
    # Crear pestaña para usuarios SharePoint system
    ws_sharepoint = wb.create_sheet("Usuarios SharePoint System")
    crear_pestana_usuarios_sharepoint(ws_sharepoint, datos_usuarios_sharepoint, campos_usuarios_sharepoint)
    
    # Guardar archivo
    nombre_archivo = crear_nombre_archivo()
    script_dir = os.path.dirname(os.path.abspath(__file__))
    ruta_completa = os.path.join(script_dir, "..", "Data", "Output", nombre_archivo)
    ruta_completa = os.path.normpath(ruta_completa)
    print(f"\nGuardando archivo: {ruta_completa}...")
    wb.save(ruta_completa)
    
    print("\n" + "=" * 80)
    print("ARCHIVO EXCEL CON DOS PESTANAS CREADO EXITOSAMENTE")
    print("=" * 80)
    print(f"Nombre: {nombre_archivo}")
    print(f"Pestana 1 - Usuarios Normales: {len(datos_usuarios_normales)} registros, {len(campos_usuarios_normales)} columnas")
    print(f"Pestana 2 - Usuarios SharePoint: {len(datos_usuarios_sharepoint)} registros, {len(campos_usuarios_sharepoint)} columnas")
    print(f"Ubicacion: Directorio Output")
    print("=" * 80)
    
    return nombre_archivo

def crear_pestana_usuarios_normales(ws, datos_registros, campos_ordenados):
    """
    Crea la pestaña para usuarios normales
    
    Args:
        ws: Worksheet de Excel
        datos_registros (list): Datos de usuarios normales
        campos_ordenados (list): Campos ordenados
    """
    print("\nCreando pestana 'Usuarios Normales'...")
    
    # Definir estilos
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    border_side = Side(style='thin', color='000000')
    header_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # Organizar columnas por bloques
    campos_base = ['RecordId', 'CreationDate', 'RecordType', 'Operation', 'UserId']
    campos_finales = ['AssociatedAdminUnits', 'AssociatedAdminUnitsNames']
    campos_app_context = sorted([c for c in campos_ordenados if 'AppAccessContext_' in c])
    campos_json = sorted([c for c in campos_ordenados 
                          if c not in campos_base 
                          and c not in campos_finales 
                          and 'AppAccessContext_' not in c])
    
    # Ordenar headers: base + app_context + json + finales
    headers_ordenados = campos_base + campos_app_context + campos_json + campos_finales
    
    # Crear headers
    for col_idx, header in enumerate(headers_ordenados, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border
    
    # Llenar datos
    for row_idx, registro in enumerate(datos_registros, 2):
        for col_idx, campo in enumerate(headers_ordenados, 1):
            valor = registro.get(campo, 'N/A')
            ws.cell(row=row_idx, column=col_idx, value=valor)
    
    # Ajustar ancho de columnas
    for col_idx, campo in enumerate(headers_ordenados, 1):
        column_letter = ws.cell(row=1, column=col_idx).column_letter
        max_length = max(
            len(str(campo)),
            max(len(str(ws.cell(row=row_idx, column=col_idx).value or '')) 
                for row_idx in range(1, len(datos_registros) + 2))
        )
        adjusted_width = min(max(max_length + 2, 12), 60)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    print(f"   Pestana 'Usuarios Normales' creada: {len(datos_registros)} filas, {len(headers_ordenados)} columnas")

def crear_pestana_usuarios_sharepoint(ws, datos_registros, campos_ordenados):
    """
    Crea la pestaña para usuarios SharePoint system
    
    Args:
        ws: Worksheet de Excel
        datos_registros (list): Datos de usuarios SharePoint
        campos_ordenados (list): Campos ordenados
    """
    print("\nCreando pestana 'Usuarios SharePoint System'...")
    
    # Definir estilos
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid')  # Rojo para diferenciar
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    border_side = Side(style='thin', color='000000')
    header_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # TODAS LAS COLUMNAS SHAREPOINT: base + todos los campos JSON + finales
    campos_base = ['RecordId', 'CreationDate', 'RecordType', 'Operation', 'UserId']
    campos_finales = ['AssociatedAdminUnits', 'AssociatedAdminUnitsNames']
    
    # Campos del JSON SharePoint (ordenados alfabéticamente)
    campos_json_sharepoint = sorted([c for c in campos_ordenados 
                                     if c not in campos_base 
                                     and c not in campos_finales])
    
    # Ordenar headers: base + json + finales
    headers_ordenados = campos_base + campos_json_sharepoint + campos_finales
    
    # Crear headers
    for col_idx, header in enumerate(headers_ordenados, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = header_border
    
    # Llenar datos
    for row_idx, registro in enumerate(datos_registros, 2):
        for col_idx, campo in enumerate(headers_ordenados, 1):
            valor = registro.get(campo, 'N/A')
            ws.cell(row=row_idx, column=col_idx, value=valor)
    
    # Ajustar ancho de columnas
    for col_idx, campo in enumerate(headers_ordenados, 1):
        column_letter = ws.cell(row=1, column=col_idx).column_letter
        max_length = max(
            len(str(campo)),
            max(len(str(ws.cell(row=row_idx, column=col_idx).value or '')) 
                for row_idx in range(1, len(datos_registros) + 2))
        )
        adjusted_width = min(max(max_length + 2, 12), 60)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    print(f"   Pestana 'Usuarios SharePoint System' creada: {len(datos_registros)} filas, {len(headers_ordenados)} columnas")

def crear_excel_completo(datos_registros, campos_ordenados):
    """
    Crea un archivo Excel con TODAS las columnas (incluido JSON aplanado)
    
    Args:
        datos_registros (list): Lista de diccionarios con todos los datos
        campos_ordenados (list): Lista ordenada de campos (columnas)
        
    Returns:
        str: Nombre del archivo creado
    """
    print("\n📊 CREANDO ARCHIVO EXCEL COMPLETO...")
    print("=" * 80)
    
    # Crear workbook y worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos Purview Completos"
    
    # Definir estilos
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    border_side = Side(style='thin', color='000000')
    header_border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # Organizar columnas por bloques para mejor visualización
    print("\n📋 Organizando columnas por bloques...")
    
    # BLOQUE 1: Campos base (siempre al inicio)
    campos_base = ['RecordId', 'CreationDate', 'RecordType', 'Operation', 'UserId']
    
    # BLOQUE 2: Campos de AppAccessContext
    campos_app_context = sorted([c for c in campos_ordenados if 'AppAccessContext_' in c])
    
    # BLOQUE 3: Campos del JSON (resto)
    campos_json = sorted([c for c in campos_ordenados 
                          if c not in campos_base 
                          and 'AppAccessContext_' not in c
                          and c not in ['AssociatedAdminUnits', 'AssociatedAdminUnitsNames']])
    
    # BLOQUE 4: Campos finales
    campos_finales = ['AssociatedAdminUnits', 'AssociatedAdminUnitsNames']
    
    # Orden final: BASE + APP_CONTEXT + JSON + FINALES
    headers_ordenados = []
    headers_ordenados.extend([c for c in campos_base if c in campos_ordenados])
    headers_ordenados.extend(campos_app_context)
    headers_ordenados.extend(campos_json)
    headers_ordenados.extend([c for c in campos_finales if c in campos_ordenados])
    
    print(f"   ✓ Bloque 1 (Base): {len([c for c in campos_base if c in campos_ordenados])} campos")
    print(f"   ✓ Bloque 2 (AppAccessContext): {len(campos_app_context)} campos")
    print(f"   ✓ Bloque 3 (JSON AuditData): {len(campos_json)} campos")
    print(f"   ✓ Bloque 4 (Finales): {len([c for c in campos_finales if c in campos_ordenados])} campos")
    print(f"   ✓ TOTAL: {len(headers_ordenados)} columnas")
    
    # Agregar headers con formato
    print("\n📝 Escribiendo headers...")
    for col, header in enumerate(headers_ordenados, 1):
        celda = ws.cell(row=1, column=col, value=header)
        celda.font = header_font
        celda.fill = header_fill
        celda.alignment = header_alignment
        celda.border = header_border
    
    # Congelar primera fila (headers)
    ws.freeze_panes = 'A2'
    
    # Agregar datos fila por fila
    print(f"\n📊 Escribiendo {len(datos_registros)} filas de datos...")
    for fila_num, registro in enumerate(datos_registros, 2):  # Empezar en fila 2
        for col, campo in enumerate(headers_ordenados, 1):
            valor = registro.get(campo, 'N/A')
            
            # Convertir valores booleanos y None
            if isinstance(valor, bool):
                valor = 'Sí' if valor else 'No'
            elif valor is None:
                valor = 'N/A'
            
            ws.cell(row=fila_num, column=col, value=str(valor))
        
        if fila_num % 100 == 0:
            print(f"   ✓ Procesadas {fila_num - 1} filas...")
    
    # Ajustar ancho de columnas
    print("\n🔧 Ajustando ancho de columnas...")
    for col_idx, column in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = column[0].column_letter
        
        for cell in column:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        
        # Ancho: mínimo 12, máximo 60
        adjusted_width = min(max(max_length + 2, 12), 60)
        ws.column_dimensions[column_letter].width = adjusted_width
        
        if col_idx % 10 == 0:
            print(f"   ✓ Ajustadas {col_idx} columnas...")
    
    # Guardar archivo en carpeta Output
    nombre_archivo = crear_nombre_archivo()
    script_dir = os.path.dirname(os.path.abspath(__file__))
    ruta_completa = os.path.join(script_dir, "..", "Data", "Output", nombre_archivo)
    ruta_completa = os.path.normpath(ruta_completa)
    print(f"\n💾 Guardando archivo: {ruta_completa}...")
    wb.save(ruta_completa)
    
    print("\n" + "=" * 80)
    print("✅ ARCHIVO EXCEL CREADO EXITOSAMENTE")
    print("=" * 80)
    print(f"📄 Nombre: {nombre_archivo}")
    print(f"📊 Filas de datos: {len(datos_registros)}")
    print(f"📋 Columnas totales: {len(headers_ordenados)}")
    print(f"💾 Ubicación: Directorio actual")
    print("=" * 80)
    
    return nombre_archivo

def crear_resumen_columnas(campos_usuarios_normales, campos_usuarios_sharepoint, nombre_archivo_base="resumen_columnas.txt"):
    """
    Crea un archivo de texto con el resumen de todas las columnas de ambos tipos
    
    Args:
        campos_usuarios_normales (list): Lista de campos de usuarios normales
        campos_usuarios_sharepoint (list): Lista de campos de usuarios SharePoint
        nombre_archivo_base (str): Nombre del archivo de resumen
    """
    ahora = datetime.now()
    timestamp = ahora.strftime("%d%m%Y_%H%M")
    nombre_archivo = f"resumen_columnas_{timestamp}.txt"
    script_dir = os.path.dirname(os.path.abspath(__file__))
    ruta_completa = os.path.join(script_dir, "..", "Data", "Output", nombre_archivo)
    ruta_completa = os.path.normpath(ruta_completa)
    
    with open(ruta_completa, 'w', encoding='utf-8') as f:
        f.write("=" * 80 + "\n")
        f.write("RESUMEN DE COLUMNAS DEL EXCEL GENERADO - DOS PESTANAS\n")
        f.write("=" * 80 + "\n\n")
        
        f.write(f"Fecha: {ahora.strftime('%d/%m/%Y %H:%M:%S')}\n")
        f.write(f"Total columnas usuarios normales: {len(campos_usuarios_normales)}\n")
        f.write(f"Total columnas usuarios SharePoint: {len(campos_usuarios_sharepoint)}\n\n")
        
        # Resumen usuarios normales
        f.write("PESTANA 1 - USUARIOS NORMALES:\n")
        f.write("-" * 40 + "\n")
        for i, campo in enumerate(sorted(campos_usuarios_normales), 1):
            f.write(f"{i:3d}. {campo}\n")
        
        f.write(f"\nPESTANA 2 - USUARIOS SHAREPOINT SYSTEM:\n")
        f.write("-" * 40 + "\n")
        for i, campo in enumerate(sorted(campos_usuarios_sharepoint), 1):
            f.write(f"{i:3d}. {campo}\n")
        
        f.write("\n" + "=" * 80 + "\n")
        f.write("FIN DEL RESUMEN\n")
        f.write("=" * 80 + "\n")
    
    print(f"\nResumen guardado en: {nombre_archivo}")
    return nombre_archivo

