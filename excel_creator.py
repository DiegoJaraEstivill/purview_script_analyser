from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

def crear_nombre_archivo():
    """Crea el nombre del archivo Excel con formato PurviewInf_DDMMAAAA_HHMM"""
    ahora = datetime.now()
    fecha = ahora.strftime("%d%m%Y")  # día-mes-año
    hora = ahora.strftime("%H%M")     # hora-minuto
    return f"PurviewInf_{fecha}_{hora}.xlsx"

def crear_excel_purview(datos_registros):
    """
    Crea un archivo Excel con los datos de los registros de Purview
    
    Args:
        datos_registros (list): Lista de diccionarios con los datos de cada registro
        
    Returns:
        str: Nombre del archivo creado
    """
    # Crear workbook y worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Datos Purview"
    
    # Definir headers para las 15 columnas (5 originales + 10 del audit JSON)
    headers = [
        'RecordID',              # A
        'Creation Date',         # B
        'Record Type',           # C
        'Operation',             # D
        'User ID',              # E
        'Audit Creation Time',   # F
        'Audit ID',             # G
        'Audit Operation',       # H
        'Organization ID',       # I
        'Audit Record Type',     # J
        'User Key',             # K
        'User Type',            # L
        'Version',              # M
        'Workload',             # N
        'Client IP'             # O
    ]
    
    # Agregar headers con formato
    for col, header in enumerate(headers, 1):
        celda = ws.cell(row=1, column=col, value=header)
        celda.font = Font(bold=True, color='FFFFFF')
        celda.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    
    # Agregar datos fila por fila
    for row, registro in enumerate(datos_registros, 2):  # Empezar en fila 2 (después del header)
        ws.cell(row=row, column=1, value=registro.get('record_id', 'N/A'))
        ws.cell(row=row, column=2, value=registro.get('creation_date', 'N/A'))
        ws.cell(row=row, column=3, value=registro.get('record_type', 'N/A'))
        ws.cell(row=row, column=4, value=registro.get('operation', 'N/A'))
        ws.cell(row=row, column=5, value=registro.get('user_id', 'N/A'))
        ws.cell(row=row, column=6, value=registro.get('audit_creation_time', 'N/A'))
        ws.cell(row=row, column=7, value=registro.get('audit_id', 'N/A'))
        ws.cell(row=row, column=8, value=registro.get('audit_operation', 'N/A'))
        ws.cell(row=row, column=9, value=registro.get('organization_id', 'N/A'))
        ws.cell(row=row, column=10, value=registro.get('audit_record_type', 'N/A'))
        ws.cell(row=row, column=11, value=registro.get('user_key', 'N/A'))
        ws.cell(row=row, column=12, value=registro.get('user_type', 'N/A'))
        ws.cell(row=row, column=13, value=registro.get('version', 'N/A'))
        ws.cell(row=row, column=14, value=registro.get('workload', 'N/A'))
        ws.cell(row=row, column=15, value=registro.get('client_ip', 'N/A'))
    
    # Ajustar ancho de columnas automáticamente
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        # Establecer un ancho mínimo de 12 y máximo de 50
        adjusted_width = min(max(max_length + 2, 12), 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Guardar archivo con nombre generado automáticamente
    nombre_archivo = crear_nombre_archivo()
    wb.save(nombre_archivo)
    
    print(f"✅ Archivo Excel creado exitosamente: {nombre_archivo}")
    print(f"📊 Total de registros procesados: {len(datos_registros)}")
    
    return nombre_archivo